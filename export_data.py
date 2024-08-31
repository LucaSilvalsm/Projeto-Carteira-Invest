from flask import current_app as app
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from Model import db, Ativos, Rendimentos, Estudos, Preco_teto
from Controller.PrecoController import PrecoController
from Controller.EstudosController import EstudosController
from Controller.UsuarioController import UsuarioController

def criar_estilos():
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    estilo_real = NamedStyle(name='estilo_real', number_format='"R$ " #,##0.00')
    estilo_percentagem = NamedStyle(name='estilo_percentagem', number_format=FORMAT_PERCENTAGE)
    return border, center_alignment, estilo_real, estilo_percentagem

def aplicar_formatacao(worksheet, border, center_alignment, estilo_real, estilo_percentagem):
    colunas_reais = ['media_recebida', 'preco_pessoal', 'valor_recebido', 'preco_medio', 'media_recebida','preco_teto']
    colunas_percentagem = ['media_rendimento', 'media_dividendos', 'media_rendimentos']
    col_names = {cell.value: cell.column for cell in worksheet[1]}
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = center_alignment
    for col_name, col_idx in col_names.items():
        if col_name in colunas_reais:
            for cell in worksheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    c.style = estilo_real
                    c.border = border
                    c.alignment = center_alignment
        if col_name in colunas_percentagem:
            for cell in worksheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    c.style = estilo_percentagem
                    c.border = border
                    c.alignment = center_alignment

def ajustar_largura_colunas(worksheet, largura_pixels=139):
    largura_coluna = largura_pixels / 7.5
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(max_length, largura_coluna)
        worksheet.column_dimensions[column].width = adjusted_width

def exportar_dados_para_excel(usuario_id):
    try:
        with app.app_context():
            session = db.session
            
            # Verifique se a conexão está estabelecida
            if session.bind is None:
                session.bind = db.engine
            
            # Consultas com SQLAlchemy
            ativos_query = session.query(Ativos).filter(Ativos.usuario_id == usuario_id)
            rendimentos_query = session.query(Rendimentos).filter(Rendimentos.usuario_id == usuario_id)
            estudos_query = session.query(Estudos).filter(Estudos.usuario_id == usuario_id)
            precos_teto_query = session.query(Preco_teto).filter(Preco_teto.usuario_id == usuario_id)

            # Certifique-se de que `statement` está acessível e não é nulo
            ativos_df = pd.read_sql(ativos_query.statement, session.bind)
            rendimentos_df = pd.read_sql(rendimentos_query.statement, session.bind)
            estudos_df = pd.read_sql(estudos_query.statement, session.bind)
            precos_teto_df = pd.read_sql(precos_teto_query.statement, session.bind)

            # Adiciona logs para verificar o conteúdo dos DataFrames
         
            
            precoController = PrecoController()
            estudosController = EstudosController()
            estudos_preco_teto_dados = estudosController.calcular_preco_teto_por_usuario(usuario_id)
            
            preco_teto = precoController.calcular_preco_teto_por_usuario(usuario_id)

            # Atualiza o DataFrame com os valores de preco_teto
            precos_teto_df['preco_teto'] = precos_teto_df['ticket'].map(lambda ticket: preco_teto.get(ticket, {}).get('preco_teto_calculado'))
            estudos_df['preco_teto'] = estudos_df['ticket_escolhido'].map(lambda ticket_escolhido: estudos_preco_teto_dados.get(ticket_escolhido, {}).get('preco_teto_calculado'))

            usuarioController = UsuarioController()
            nome_usuario = usuarioController.obter_nome_usuario(usuario_id)
            

            # Adiciona um log para verificar o DataFrame depois de adicionar preco_teto
          
            
            arquivo_excel = f'Dados_Carteira_{usuario_id}_{nome_usuario}.xlsx'
            with pd.ExcelWriter(arquivo_excel, engine='openpyxl') as writer:
                ativos_df.to_excel(writer, sheet_name='Ativos', index=False)
                rendimentos_df.to_excel(writer, sheet_name='Rendimentos', index=False)
                estudos_df.to_excel(writer, sheet_name='Estudos', index=False)
                precos_teto_df.to_excel(writer, sheet_name='Preços Teto', index=False)

            wb = load_workbook(arquivo_excel)
            border, center_alignment, estilo_real, estilo_percentagem = criar_estilos()
            for sheet_name in wb.sheetnames:
                worksheet = wb[sheet_name]
                aplicar_formatacao(worksheet, border, center_alignment, estilo_real, estilo_percentagem)
                ajustar_largura_colunas(worksheet)
            wb.save(arquivo_excel)

            print(f"Dados exportados e formatados para '{arquivo_excel}'")
            return arquivo_excel
    except Exception as e:
        print(f"Erro ao exportar dados para o Excel: {e}")
        return None

